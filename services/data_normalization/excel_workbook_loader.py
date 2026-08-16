# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025-2026 Sönke Pinkernelle

# PortfoliFLOW — Excel import workbook loader
"""Excel import workbook loader — canonical parsing path (ADR-0009).

Purpose and responsibility:
    This module is the single authoritative parser for the multi-sheet
    PortfoliFLOW Excel import workbook. It converts an ``.xlsx`` file into the
    set of canonical DataFrames consumed by the downstream investment
    extractor and every other module.

Layering:
    ``docs/architecture.md`` names ``services/data_normalization/`` as the
    shared parsing path. This module lives here so that both the web surface
    (``web/routes/data_import.py``) and the Front-Office module shell can
    consume it without ``web/`` importing from ``modules/`` (the hard layering
    rule in ``CLAUDE.md``). It imports only stdlib, third-party (``pandas``,
    ``openpyxl``), and ``core/`` — never from ``modules/`` or ``web/``.

    The public entry point :func:`load_excel` is re-exported from
    ``modules.front_office.data_import`` for backward-compatible imports
    pending test migration.

Supported input formats:
    * Excel (.xlsx) via ``openpyxl``.  Future: CSV, Bloomberg, Reuters.

Expected workbook layout (per ADR-0009):
    The workbook contains the named sheets listed in ``RECOGNIZED_SHEETS``.
    Every sheet shares the same first three rows:

    * **Row 1**: Column A is empty; columns B onward contain investment names
      (e.g. ``Investition A``, ``Investition B``, ...).  The number of
      investment columns is discovered dynamically by scanning row 1 — it is
      never hardcoded.
    * **Row 2**: Investment type labels (e.g. ``Aktien``, ``Private Equity``).
      Placeholder columns may contain ``Typ der Investition``.
    * **Row 3**: Investment sub-class labels.  Placeholder columns may contain
      ``Klasse der Investition``.

    The ``Attributes`` sheet has key-value attribute rows from row 4 onward.
    Investment time-series sheets share the same column namespace (investment
    names) and have date-indexed numeric rows from row 4 onward.
    Market reference sheets (e.g. ``interest rates``) have their own independent
    column namespaces (rate/index names) and the same date-indexed row structure.

Canonical output:
    :func:`load_excel` returns ``dict[str, pd.DataFrame]`` keyed by the
    snake_case sheet name (e.g. ``"navs_actual"``, ``"attributes"``).
"""

from __future__ import annotations

import datetime
import logging
import pathlib
from typing import Any

import pandas as pd

from core.exceptions import DataImportError, ValidationError

logger = logging.getLogger(__name__)

__all__ = ["load_excel", "validate_dataframe", "validate_workbook"]

# ---------------------------------------------------------------------------
# Sheet categories — every recognised sheet belongs to exactly one category.
#
# Investment time-series sheets share the same column namespace (investment
# names) and participate in cross-sheet column consistency validation.
#
# Market reference sheets have their own independent column namespaces
# (rate or index names).  They are excluded from the investment column
# consistency check.  More sheets may be added to MARKET_REFERENCE_SHEETS
# in the future without touching the parsing logic.
# ---------------------------------------------------------------------------

ATTRIBUTES_SHEET: str = "Attributes"

INVESTMENT_TIMESERIES_SHEETS: frozenset[str] = frozenset(
    {
        "Cash Flow In actual",
        "Cash Flow In plan",
        "Cash Flow Out actual",
        "Cash Flow Out plan",
        # Liquid-archetype income sheets (ADR-0081). Same wide,
        # investment-keyed idiom as the other cash-flow sheets, so they
        # share the investment-column namespace and participate in the
        # cross-sheet consistency check. ``flow_type`` is derived from
        # the resolved investment type at the extractor layer.
        "Cash Flow Income actual",
        "Cash Flow Income plan",
        "NAVs actual",
        "NAVs plan",
        # Cash statement sheet (ADR-0103 §3, workbook v32). The
        # statement-style **book of record** for cash balances: one column
        # per cash position, one row per statement date, each cell the
        # balance in the position's currency (a *level*, not a flow — the
        # importer derives the ledger deltas from it, ADR-0103 §4). The
        # v31 ``Cash USD`` NAV column moves here.
        #
        # Registered as an ordinary wide investment time-series sheet: it
        # shares the investment-column namespace and participates in the
        # cross-sheet consistency check, so it carries the *same* column
        # headers as every other investment sheet — the non-cash columns
        # simply stay empty (empty columns are valid placeholder slots,
        # ADR-0009). The unchanged wide parser handles it, which is the
        # entire loader delta.
        #
        # The sheet is optional: a workbook without it (v31) parses
        # byte-identically, and the extractor keeps the v31 cash-NAV-column
        # path alive for that case.
        "Cash",
        "total return actual",
        "total return plan",
    }
)

MARKET_REFERENCE_SHEETS: frozenset[str] = frozenset(
    {
        "interest rates",
        "AUM",
        "Benchmarks actual",
        # FX-rate series (ADR-0099 §5). Wide-format market-reference
        # sheet exactly like ``Benchmarks actual`` (ADR-0061): column A
        # = dates, one column per rate series. The column headers use
        # pair notation (``USD/EUR``, ``GBP/EUR``); the standard
        # market-reference parser reads them verbatim and the FX
        # extractor validates the pair convention. Canonical key
        # auto-derives to ``fx_rates``. The sheet is optional — a
        # workbook without it imports byte-identically.
        "FX rates",
    }
)

# Phase-7 Anlagegrenzen-Überwachung sheets (ADR-0056). The two limit-set
# sheets carry a fundamentally different row schema (per-set columns,
# effective_from / label / notes / class-key rows) and are parsed by
# ``_parse_limit_set_sheet``. Each sheet has its own independent column
# namespace and is excluded from the investment column-consistency check.
LIMIT_SET_SHEETS: frozenset[str] = frozenset(
    {
        "Limit Set SAA",
        "Limit Set 2",
    }
)

# Phase-7 Benchmarks & Attribution sheets (ADR-0061). The mapping sheet
# carries (asset_class, benchmark_id, weight, comment) rows and is
# parsed by ``_parse_benchmark_mapping_sheet``. The companion
# ``Benchmarks actual`` time-series sheet is registered as a
# market-reference sheet above and inherits the standard parser.
BENCHMARK_MAPPING_SHEETS: frozenset[str] = frozenset(
    {
        "Benchmark Mapping",
    }
)

# Liquid-archetype reference sheets (ADR-0081 / ADR-0079). Each is a
# tidy/long sheet (one row per natural-key tuple) with its own column
# schema, parsed by a dedicated parser on the ``Benchmark Mapping``
# pattern. They do **not** share the investment-name column namespace
# and are excluded from the cross-sheet investment-column-consistency
# check.
LIQUID_REFERENCE_SHEETS: frozenset[str] = frozenset(
    {
        "Bond Analytics",
        "Rating Weights",
        "Maturity Weights",
    }
)

RECOGNIZED_SHEETS: frozenset[str] = (
    {ATTRIBUTES_SHEET}
    | INVESTMENT_TIMESERIES_SHEETS
    | MARKET_REFERENCE_SHEETS
    | LIMIT_SET_SHEETS
    | BENCHMARK_MAPPING_SHEETS
    | LIQUID_REFERENCE_SHEETS
)

# Canonical snake_case keys for market reference sheets.
# Derived from MARKET_REFERENCE_SHEETS using the same transformation as
# _sheet_name_to_key, kept in sync automatically.
_MARKET_REFERENCE_KEYS: frozenset[str] = frozenset(
    s.lower().replace(" ", "_") for s in MARKET_REFERENCE_SHEETS
)

# Canonical snake_case keys for limit-set sheets. Same derivation rule.
_LIMIT_SET_KEYS: frozenset[str] = frozenset(s.lower().replace(" ", "_") for s in LIMIT_SET_SHEETS)

# Canonical snake_case keys for benchmark-mapping sheets. Excluded
# from the investment-column-consistency check because the column
# namespace differs (the four operator-facing columns
# ``asset_class | benchmark_id | weight | comment``).
_BENCHMARK_MAPPING_KEYS: frozenset[str] = frozenset(
    s.lower().replace(" ", "_") for s in BENCHMARK_MAPPING_SHEETS
)

# Canonical snake_case keys for liquid-archetype reference sheets.
# Same derivation rule; excluded from the investment-column-consistency
# check because each carries its own tidy/long column schema.
_LIQUID_REFERENCE_KEYS: frozenset[str] = frozenset(
    s.lower().replace(" ", "_") for s in LIQUID_REFERENCE_SHEETS
)


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _sheet_name_to_key(sheet_name: str) -> str:
    """Derive a snake_case DataStore key from a workbook sheet name.

    The transformation is computed (not looked up), so new sheet names
    map automatically without code changes.

    Args:
        sheet_name: Raw sheet name from the workbook (e.g. ``"NAVs actual"``).

    Returns:
        Snake_case string suitable as a dict key (e.g. ``"navs_actual"``).
    """
    return sheet_name.lower().replace(" ", "_")


def _discover_investment_columns(ws: Any) -> list[str]:
    """Read row 1 of a worksheet and return investment column names.

    Scans from column B (openpyxl 1-based index 2) rightward, collecting all
    non-empty cell values as investment names.  Stops at the first empty cell
    or at the end of the row.

    This is the **sole** mechanism for determining how many investments exist in
    a workbook sheet.  No column count, column letter, or index is hardcoded
    anywhere in the parser — the column span is always derived from this
    function's result.

    Args:
        ws: An openpyxl ``Worksheet`` object (any access mode).

    Returns:
        Ordered list of investment name strings, one per non-empty column in
        row 1 starting from column B.  Returns an empty list if no investment
        columns are found.
    """
    names: list[str] = []
    max_col = ws.max_column
    if max_col is None or max_col < 2:
        # Sheet has no columns beyond A — no investments
        return names
    # Column 1 (A) is always the label/date column; start at column 2 (B)
    for col_idx in range(2, max_col + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is None or str(val).strip() == "":
            # First empty cell marks the end of the investment column range
            break
        names.append(str(val).strip())
    return names


def _parse_attributes_sheet(ws: Any, investment_names: list[str]) -> pd.DataFrame:
    """Parse the ``Attributes`` sheet into a keyed-attribute DataFrame.

    Rows 2 and 3 are prepended as synthetic ``"Investment Type"`` and
    ``"Investment Sub-Class"`` rows so that all metadata about each investment
    is available from a single DataFrame.

    Args:
        ws: The openpyxl ``Worksheet`` for the ``Attributes`` sheet.
        investment_names: Investment column names discovered from row 1,
            as returned by :func:`_discover_investment_columns`.

    Returns:
        DataFrame with:

        * **Index**: attribute labels starting with ``"Investment Type"`` and
          ``"Investment Sub-Class"``, followed by all key labels found in
          column A of rows 4 onward.
        * **Columns**: investment names (all columns, including empty placeholders).
        * **Values**: raw cell values; mixed dtypes are expected.
    """
    n = len(investment_names)
    # Investment data occupies columns 2 through n+1 (openpyxl 1-indexed)
    inv_max_col = n + 1

    # Row 2: investment type labels
    type_row = [ws.cell(row=2, column=c).value for c in range(2, inv_max_col + 1)]

    # Row 3: investment sub-class labels
    subclass_row = [ws.cell(row=3, column=c).value for c in range(2, inv_max_col + 1)]

    # Rows 4+: attribute labels (column A) and their values per investment
    attr_labels: list[str] = []
    attr_data: list[list[Any]] = []
    row_idx = 4
    while True:
        label_val = ws.cell(row=row_idx, column=1).value
        if label_val is None:
            break
        attr_labels.append(str(label_val))
        values = [ws.cell(row=row_idx, column=c).value for c in range(2, inv_max_col + 1)]
        attr_data.append(values)
        row_idx += 1

    index = ["Investment Type", "Investment Sub-Class", *attr_labels]
    data_rows = [type_row, subclass_row, *attr_data]
    return pd.DataFrame(data_rows, index=index, columns=investment_names)


def _parse_timeseries_sheet(ws: Any, investment_names: list[str]) -> pd.DataFrame:
    """Parse a time-series sheet into a date-indexed float DataFrame.

    Rows 2 and 3 (type/sub-class metadata) are intentionally skipped: that
    information is already captured in the ``Attributes`` sheet.

    Empty-value cells are coerced to ``NaN``; entirely-empty plan sheets
    (where all value cells are ``None`` or non-numeric) are valid and return
    an all-NaN DataFrame.

    Args:
        ws: The openpyxl ``Worksheet`` for a time-series sheet.
        investment_names: Investment column names from row 1.

    Returns:
        DataFrame with:

        * **Index**: ``DatetimeIndex`` named ``"Date"``, sorted chronologically.
        * **Columns**: investment names (all columns, including empty placeholders).
        * **Values**: ``float64``; ``NaN`` for missing / not-yet-started investments.
    """
    n = len(investment_names)
    inv_max_col = n + 1  # columns A (1) through last investment (n+1)

    dates: list[Any] = []
    rows_data: list[list[Any]] = []

    # iter_rows is more efficient than repeated ws.cell() calls for large sheets
    for row in ws.iter_rows(min_row=4, min_col=1, max_col=inv_max_col):
        date_cell = row[0]
        if date_cell.value is None:
            # First missing date marks end of data (continuous date range assumed)
            break
        dates.append(date_cell.value)
        # row[1:] are exactly the n investment value cells
        values = [cell.value for cell in row[1:]]
        rows_data.append(values)

    if dates:
        # Normalise timestamps to midnight (strips any sub-day precision from Excel)
        date_index = pd.DatetimeIndex(
            pd.to_datetime(dates, errors="coerce").normalize(), name="Date"
        )
        df = pd.DataFrame(rows_data, index=date_index, columns=investment_names)
    else:
        # Plan sheet with no dates at all — return an empty but well-typed DataFrame
        date_index = pd.DatetimeIndex([], name="Date")
        df = pd.DataFrame(index=date_index, columns=investment_names, dtype="float64")
        return df

    # Coerce non-numeric cells (strings, None, mixed types) to NaN, then cast
    df = df.apply(pd.to_numeric, errors="coerce").astype("float64")
    return df.sort_index()


def _discover_limit_set_columns(ws: Any) -> list[str]:
    """Read row 1 of a limit-set sheet and return per-set header labels.

    Same dynamic-discovery shape as :func:`_discover_investment_columns`:
    scan column B onward until the first empty cell. Each non-empty value
    is the display label of one limit set (e.g. ``"Set 1: SAA initial"``).

    Args:
        ws: An openpyxl ``Worksheet`` object for a limit-set sheet.

    Returns:
        Ordered list of set-label strings, one per non-empty column in
        row 1 starting from column B.
    """
    names: list[str] = []
    max_col = ws.max_column
    if max_col is None or max_col < 2:
        return names
    for col_idx in range(2, max_col + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is None or str(val).strip() == "":
            break
        names.append(str(val).strip())
    return names


def _parse_limit_set_sheet(ws: Any, set_labels: list[str]) -> pd.DataFrame:
    """Parse a limit-set sheet into a string-indexed DataFrame.

    Layout (ADR-0056 §Implementation pointers, refined for the v21
    Excel format):

    * Row 1: column A blank, columns B+ are operator-readable set
      headers (display only — discarded by the parser).
    * Row 2 column A = ``"effective_from"``, columns B+ = dates.
    * Row 3 column A = ``"label"``, columns B+ = strings.
    * Row 4 column A = ``"notes"``, columns B+ = strings.
    * Row 5: blank separator.
    * Row 6 column A = ``"Class Key"``, columns B+ blank (header).
    * Rows 7+ column A = class_key, columns B+ = limit percentages.
    * Parsing stops at the first row where column A is blank or
      starts with ``"Sum"`` (case-insensitive) — the operator-facing
      ``"Sum (check)"`` control row is ignored.

    Args:
        ws: The openpyxl ``Worksheet`` for a limit-set sheet.
        set_labels: Operator-readable set headers from row 1 (length
            = number of sets in the sheet).

    Returns:
        DataFrame with:

        * **Index**: ``"effective_from"``, ``"label"``, ``"notes"``,
          then every class_key in document order.
        * **Columns**: integer positional indices ``0..len(set_labels)-1``
          (the operator-facing labels live as the synthetic ``"label"``
          row, not as column names — set identity in the data layer
          is positional).
        * **Values**: raw cell values; mixed dtypes are expected
          (datetimes for ``effective_from``, strings for ``label`` /
          ``notes``, ``Decimal``-coercible numerics for class rows).
    """
    n_sets = len(set_labels)
    set_max_col = n_sets + 1

    effective_from = [ws.cell(row=2, column=c).value for c in range(2, set_max_col + 1)]
    label_row = [ws.cell(row=3, column=c).value for c in range(2, set_max_col + 1)]
    notes_row = [ws.cell(row=4, column=c).value for c in range(2, set_max_col + 1)]

    class_labels: list[str] = []
    class_data: list[list[Any]] = []
    row_idx = 7  # rows 5 and 6 are separator + "Class Key" header
    while True:
        label_val = ws.cell(row=row_idx, column=1).value
        if label_val is None:
            break
        label_str = str(label_val).strip()
        if not label_str:
            break
        if label_str.lower().startswith("sum"):
            # Operator-facing control row; ignore and stop (no class keys follow).
            break
        class_labels.append(label_str)
        values = [ws.cell(row=row_idx, column=c).value for c in range(2, set_max_col + 1)]
        class_data.append(values)
        row_idx += 1

    index = ["effective_from", "label", "notes", *class_labels]
    data_rows = [effective_from, label_row, notes_row, *class_data]
    # Columns are positional indices — the operator labels live in the
    # synthetic ``"label"`` data row. This keeps cross-row alignment
    # explicit and makes the per-set lookup trivial downstream.
    return pd.DataFrame(data_rows, index=index, columns=list(range(n_sets)))


_BENCHMARK_MAPPING_COLUMNS: tuple[str, ...] = (
    "asset_class",
    "benchmark_id",
    "weight",
    "comment",
)


def _parse_benchmark_mapping_sheet(ws: Any) -> pd.DataFrame:
    """Parse the ``Benchmark Mapping`` sheet into a four-column DataFrame.

    Layout (per ADR-0061 §Decision):

    * Row 1: header — ``asset_class | benchmark_id | weight | comment``.
    * Rows 2+: data — one mapping per row. Empty rows are dropped.

    The parser is intentionally schema-strict: it always returns a
    DataFrame with exactly the four canonical columns, even when the
    sheet is present but empty. The downstream extractor and service
    layer rely on the column shape for hard-fail validation.

    Args:
        ws: An openpyxl ``Worksheet`` object for the ``Benchmark Mapping``
            sheet.

    Returns:
        DataFrame with columns ``("asset_class", "benchmark_id",
        "weight", "comment")``. Empty body when the sheet has no
        data rows.
    """
    rows: list[list[Any]] = []
    # ``ws.max_row`` includes empty trailing rows for some openpyxl
    # input flavours; we iterate and drop blank rows defensively.
    max_row = ws.max_row or 1
    for row_idx in range(2, max_row + 1):
        asset_class = ws.cell(row=row_idx, column=1).value
        benchmark_id = ws.cell(row=row_idx, column=2).value
        weight = ws.cell(row=row_idx, column=3).value
        comment = ws.cell(row=row_idx, column=4).value
        # Drop entirely-empty rows; otherwise preserve the row even
        # when ``benchmark_id`` is blank (the "no benchmark for this
        # asset class" case is valid input, e.g. Cash).
        if asset_class is None and benchmark_id is None and weight is None and comment is None:
            continue
        rows.append([asset_class, benchmark_id, weight, comment])
    return pd.DataFrame(rows, columns=list(_BENCHMARK_MAPPING_COLUMNS))


# ---------------------------------------------------------------------------
# Liquid-archetype reference-sheet parsers (ADR-0081)
#
# Three tidy/long sheets, parsed on the ``_parse_benchmark_mapping_sheet``
# pattern: a single header row in row 1, data from row 2, one row per
# natural-key tuple. The parsers are intentionally lenient (blank cells
# survive as ``None``) and do **not** enforce bucket / taxonomy validity
# — that is the extractor's row-level, partial-success concern.
# ---------------------------------------------------------------------------


_BOND_ANALYTICS_COLUMNS: tuple[str, ...] = (
    "as_of_date",
    "investment",
    "ytm",
    "eff_duration",
    "oas",
    "convexity",
)
_RATING_WEIGHTS_COLUMNS: tuple[str, ...] = (
    "as_of_date",
    "investment",
    "rating_bucket",
    "weight_pct",
)
_MATURITY_WEIGHTS_COLUMNS: tuple[str, ...] = (
    "as_of_date",
    "investment",
    "maturity_bucket",
    "weight_pct",
)


def _coerce_cell_to_iso_date_str(value: Any) -> Any:
    """Coerce a date/datetime cell to an ISO date string.

    openpyxl returns ``datetime.datetime`` for date cells; serialising
    the parsed DataFrame via ``DataFrame.to_json`` would otherwise leave
    the encoding to pandas dtype inference. Normalising to a plain ISO
    string here keeps the JSONB round-trip deterministic and matches the
    string-date shape the extractor's ``_parse_iso_date`` consumes.

    Non-date values are returned unchanged (the extractor flags an
    unparseable date as a row-level error).
    """
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _parse_tidy_reference_sheet(ws: Any, columns: tuple[str, ...]) -> pd.DataFrame:
    """Parse a tidy/long reference sheet into a fixed-column DataFrame.

    Reads columns positionally (column A → ``columns[0]``, etc.) from
    row 2 onward, dropping entirely-empty rows. The ``as_of_date``
    column is normalised to an ISO date string; every other cell is kept
    verbatim (numerics stay numeric, blanks stay ``None``). The returned
    DataFrame always carries exactly ``columns`` so the downstream
    extractor can rely on the column shape.

    Args:
        ws: An openpyxl ``Worksheet`` for a liquid-reference sheet.
        columns: The canonical column names, in positional order.

    Returns:
        DataFrame with a plain ``RangeIndex`` and the canonical columns;
        empty body when the sheet has no data rows.
    """
    n_cols = len(columns)
    date_positions = {idx for idx, name in enumerate(columns) if name == "as_of_date"}
    rows: list[list[Any]] = []
    max_row = ws.max_row or 1
    for row_idx in range(2, max_row + 1):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, n_cols + 1)]
        if all(cell is None for cell in cells):
            continue
        coerced = [
            _coerce_cell_to_iso_date_str(cell) if pos in date_positions else cell
            for pos, cell in enumerate(cells)
        ]
        rows.append(coerced)
    return pd.DataFrame(rows, columns=list(columns))


_LIQUID_REFERENCE_COLUMNS: dict[str, tuple[str, ...]] = {
    "Bond Analytics": _BOND_ANALYTICS_COLUMNS,
    "Rating Weights": _RATING_WEIGHTS_COLUMNS,
    "Maturity Weights": _MATURITY_WEIGHTS_COLUMNS,
}


# ---------------------------------------------------------------------------
# Public API — importable without the module or web layer
# ---------------------------------------------------------------------------


def load_excel(
    path: str | pathlib.Path,
    *,
    sheets: list[str] | None = None,
) -> dict[str, pd.DataFrame]:
    """Load a multi-sheet PortfoliFLOW Excel import workbook (per ADR-0009).

    Discovers the number of investment columns dynamically from row 1 of each
    sheet.  Files with any number of investments are supported without code
    changes — simply add columns to the Excel file and re-import.

    Args:
        path: Absolute or relative path to the ``.xlsx`` file.
        sheets: If provided, load only the listed sheet names.  Unrecognised
            names and names absent from the workbook emit WARNING log entries
            but do not raise.  If ``None`` (default), all recognised sheets
            present in the workbook are loaded.

    Returns:
        Dict mapping canonical snake_case keys to DataFrames.  For example:

        * ``"attributes"`` → key-value attribute DataFrame.
        * ``"navs_actual"`` → date-indexed NAV time-series DataFrame.
        * ``"total_return_plan"`` → date-indexed return plan DataFrame (may be
          all-NaN if not yet filled in).

    Raises:
        DataImportError: If ``path`` does not exist, if the file cannot be
            parsed as an Excel workbook, or if no recognised sheets are found.
        ValidationError: If a recognised sheet fails structural validation
            (e.g. columns mismatch across sheets).
    """
    path = pathlib.Path(path)

    # Guard: file must exist before attempting any I/O
    if not path.exists() or not path.is_file():
        raise DataImportError(f"File not found or not a regular file: {path}")

    # Guard: openpyxl is a required runtime dependency for Excel support
    try:
        import openpyxl as _openpyxl
    except ImportError as exc:
        raise DataImportError(
            "openpyxl is required for Excel import.  Install it with: pip install 'openpyxl>=3.1'"
        ) from exc

    try:
        # data_only=True: return computed cell values rather than formula strings
        wb = _openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise DataImportError(f"Cannot parse '{path.name}': {type(exc).__name__}: {exc}") from exc

    workbook_sheets = set(wb.sheetnames)

    # ------------------------------------------------------------------
    # Resolve the set of sheets to actually load
    # ------------------------------------------------------------------
    if sheets is not None:
        for requested in sheets:
            if requested not in workbook_sheets:
                logger.warning("Requested sheet '%s' not found in '%s'.", requested, path.name)
        target_sheets = [s for s in sheets if s in workbook_sheets and s in RECOGNIZED_SHEETS]
    else:
        # Warn about sheets in the file that the parser does not know about
        for name in wb.sheetnames:
            if name not in RECOGNIZED_SHEETS:
                logger.warning(
                    "Sheet '%s' in '%s' is not a recognised import-format sheet name — skipping.",
                    name,
                    path.name,
                )
        target_sheets = [s for s in wb.sheetnames if s in RECOGNIZED_SHEETS]

    if not target_sheets:
        raise DataImportError(
            f"No recognised sheets found in '{path.name}'.  "
            f"Expected one or more of: {sorted(RECOGNIZED_SHEETS)}."
        )

    logger.info(
        "Loading '%s': %d recognised sheet(s) to process.",
        path.name,
        len(target_sheets),
    )

    datasets: dict[str, pd.DataFrame] = {}

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        key = _sheet_name_to_key(sheet_name)

        if sheet_name in BENCHMARK_MAPPING_SHEETS:
            # Benchmark Mapping carries a flat four-column schema that
            # does not share the investment-column namespace; parse it
            # via the dedicated branch and skip the cross-sheet
            # consistency check downstream.
            try:
                df = _parse_benchmark_mapping_sheet(ws)
            except Exception as exc:
                raise DataImportError(
                    f"Failed to parse sheet '{sheet_name}': {type(exc).__name__}: {exc}"
                ) from exc
            datasets[key] = df
            logger.info(
                "Sheet '%s' → key='%s': %d mapping row(s) parsed.",
                sheet_name,
                key,
                len(df),
            )
            continue

        if sheet_name in LIQUID_REFERENCE_SHEETS:
            # Liquid-archetype reference sheets carry their own tidy/long
            # column schema (own column namespace); parse via the
            # dedicated branch and skip the cross-sheet investment-column
            # consistency check downstream (like Benchmark Mapping).
            try:
                df = _parse_tidy_reference_sheet(ws, _LIQUID_REFERENCE_COLUMNS[sheet_name])
            except Exception as exc:
                raise DataImportError(
                    f"Failed to parse sheet '{sheet_name}': {type(exc).__name__}: {exc}"
                ) from exc
            datasets[key] = df
            logger.info(
                "Sheet '%s' → key='%s': %d reference row(s) parsed.",
                sheet_name,
                key,
                len(df),
            )
            continue

        if sheet_name in LIMIT_SET_SHEETS:
            # Limit-set sheets have their own row schema (effective_from,
            # label, notes, class_keys). Discover the per-set columns
            # from row 1 and parse via the dedicated branch.
            set_labels = _discover_limit_set_columns(ws)
            logger.info(
                "Sheet '%s' → key='%s': discovered %d limit-set column(s).",
                sheet_name,
                key,
                len(set_labels),
            )
            try:
                df = _parse_limit_set_sheet(ws, set_labels)
                validate_dataframe(df, kind="attributes")
            except ValidationError:
                raise
            except Exception as exc:
                raise DataImportError(
                    f"Failed to parse sheet '{sheet_name}': {type(exc).__name__}: {exc}"
                ) from exc
            datasets[key] = df
            logger.info("  → Limit-set: shape=%s.", df.shape)
            continue

        investment_names = _discover_investment_columns(ws)

        # Market reference sheets carry rate/index names, not investment names.
        col_kind = "rate series" if sheet_name in MARKET_REFERENCE_SHEETS else "investment"
        logger.info(
            "Sheet '%s' → key='%s': discovered %d %s column(s).",
            sheet_name,
            key,
            len(investment_names),
            col_kind,
        )

        try:
            if sheet_name == "Attributes":
                df = _parse_attributes_sheet(ws, investment_names)
                validate_dataframe(df, kind="attributes")
            else:
                df = _parse_timeseries_sheet(ws, investment_names)
                validate_dataframe(df, kind="timeseries")
        except ValidationError:
            raise  # re-raise with original context
        except Exception as exc:
            raise DataImportError(
                f"Failed to parse sheet '{sheet_name}': {type(exc).__name__}: {exc}"
            ) from exc

        # Per-sheet summary log
        if sheet_name == "Attributes":
            logger.info("  → Attributes: shape=%s.", df.shape)
        else:
            populated = int(df.notna().any(axis=0).sum()) if not df.empty else 0
            date_range_str: str
            if not df.empty and len(df.index) > 0:
                date_range_str = f"{df.index[0].date()} … {df.index[-1].date()}"
            else:
                date_range_str = "empty"
            logger.info(
                "  → shape=%s, date range=%s, %d/%d columns have data.",
                df.shape,
                date_range_str,
                populated,
                len(investment_names),
            )

        datasets[key] = df

    wb.close()

    # Cross-sheet consistency: all sheets must share the same investment columns
    if len(datasets) > 1:
        validate_workbook(datasets)

    # Report investment column count from a non-market-reference, non-limit-set
    # sheet so that rate series and limit-set columns (which have different
    # column namespaces) are not conflated.
    investment_keys = [
        k
        for k in datasets
        if k not in _MARKET_REFERENCE_KEYS
        and k not in _LIMIT_SET_KEYS
        and k not in _BENCHMARK_MAPPING_KEYS
        and k not in _LIQUID_REFERENCE_KEYS
    ]
    n_inv_cols = len(datasets[investment_keys[0]].columns) if investment_keys else 0
    market_keys = [k for k in datasets if k in _MARKET_REFERENCE_KEYS]
    rate_summary = ", ".join(
        f"{len(datasets[k].columns)} rate series in '{k}'" for k in market_keys
    )
    if rate_summary:
        logger.info(
            "Loaded '%s': %d dataset(s), %d investment column(s), %s.",
            path.name,
            len(datasets),
            n_inv_cols,
            rate_summary,
        )
    else:
        logger.info(
            "Loaded '%s': %d dataset(s), %d investment column(s).",
            path.name,
            len(datasets),
            n_inv_cols,
        )

    return datasets


def validate_dataframe(df: pd.DataFrame, *, kind: str = "timeseries") -> None:
    """Validate that a DataFrame conforms to the PortfoliFLOW Excel import schema.

    Called automatically by :func:`load_excel` but may also be invoked directly
    when DataFrames are constructed from other sources.

    Args:
        df: The DataFrame to validate.
        kind: Schema variant to apply.  Either ``"timeseries"`` (default) or
            ``"attributes"``.  The ``"timeseries"`` variant requires a
            ``DatetimeIndex``; the ``"attributes"`` variant must NOT have one.

    Raises:
        ValidationError: With a descriptive message if any of the following
            conditions hold:

            * Fewer than one data column is present (both kinds).
            * For ``"timeseries"``: index is not a ``DatetimeIndex``, or index
              is not monotonically increasing, or any float column contains
              ``±inf`` values.
            * For ``"attributes"``: index is a ``DatetimeIndex`` (attributes
              DataFrames use string labels as their index).
            * ``kind`` is not one of the two recognised values.
    """
    # Column count applies to both schema kinds
    if df.shape[1] < 1:
        raise ValidationError(
            f"DataFrame must contain at least one data column; found {df.shape[1]}."
        )

    if kind == "timeseries":
        # Time-series schema requires a DatetimeIndex; empty plan sheets are valid
        if not isinstance(df.index, pd.DatetimeIndex):
            raise ValidationError(
                f"DataFrame index must be a DatetimeIndex, got {type(df.index).__name__}."
            )
        # Sorted check only applies when there are rows to compare
        if len(df.index) > 1 and not df.index.is_monotonic_increasing:
            raise ValidationError("DatetimeIndex must be monotonically increasing (sorted).")
        # ±inf in numeric data indicates a data-source error, not just missingness
        for col in df.columns:
            if not pd.api.types.is_float_dtype(df[col]):
                continue
            if df[col].isin([float("inf"), float("-inf")]).any():
                raise ValidationError(
                    f"Column '{col}' contains infinite values (±inf).  "
                    "Check the source data for division-by-zero artefacts."
                )

    elif kind == "attributes":
        # Attributes DataFrames use string attribute labels as their index,
        # never a DatetimeIndex
        if isinstance(df.index, pd.DatetimeIndex):
            raise ValidationError(
                "Attributes DataFrame must not have a DatetimeIndex.  "
                "The index should contain string attribute labels."
            )

    else:
        raise ValidationError(
            f"Unknown validation kind: '{kind}'.  Expected 'timeseries' or 'attributes'."
        )


def validate_workbook(datasets: dict[str, pd.DataFrame]) -> None:
    """Validate cross-sheet column consistency for investment sheets in a loaded workbook.

    All investment DataFrames in ``datasets`` must share identical column names,
    since they all represent the same set of investments.  A mismatch indicates
    that someone added a column to one sheet but not to the others.

    Market reference sheets (e.g. ``interest_rates``) are excluded from this
    check because they have their own independent column namespaces (rate/index
    names rather than investment names).  A separate consistency check for
    market reference sheets can be added when there are multiple such sheets.

    Args:
        datasets: Dict mapping canonical key → DataFrame, as returned by
            :func:`load_excel`.

    Raises:
        ValidationError: If any investment DataFrame has different column names
            from the reference investment DataFrame.
    """
    # Exclude market-reference and limit-set sheets — their columns are not
    # investment names. The investment cross-sheet consistency check only
    # applies to sheets that share the investment-name column namespace.
    investment_keys = [
        k
        for k in datasets
        if k not in _MARKET_REFERENCE_KEYS
        and k not in _LIMIT_SET_KEYS
        and k not in _BENCHMARK_MAPPING_KEYS
        and k not in _LIQUID_REFERENCE_KEYS
    ]

    if len(investment_keys) < 2:
        return  # nothing to compare among investment sheets

    reference_key = investment_keys[0]
    reference_cols = list(datasets[reference_key].columns)

    for key in investment_keys[1:]:
        cols = list(datasets[key].columns)
        if cols != reference_cols:
            raise ValidationError(
                f"Investment column mismatch: sheet '{key}' has columns {cols!r} "
                f"but '{reference_key}' has {reference_cols!r}.  "
                "All investment sheets must share the same investment columns."
            )

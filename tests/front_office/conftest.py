# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025-2026 Sönke Pinkernelle

# PortfoliFLOW — Test fixtures for the Data Import module
"""Shared pytest fixtures for ``tests/front_office/``.

All fixtures are self-contained: they generate workbooks programmatically with
``openpyxl`` and write to a ``tmp_path`` provided by pytest.  No fixture depends
on real sample files in ``data/sample/``, so the test suite runs in any CI
environment without extra data files.

Fixture overview
----------------
Excel-import-format fixtures (current format, per ADR-0009):

* ``sample_excel_import_workbook`` — 5 investment columns (3 active,
  2 placeholders), 10 daily dates.  The canonical fixture for all
  Excel-import parser tests.
* ``sample_excel_import_workbook_extended`` — 8 investment columns
  (6 active, 2 placeholders), same date range.  Used specifically to
  prove that the parser adapts to different column counts without code
  changes.

Legacy fixtures (V1 single-sheet format, kept for backward-compatibility tests):

* ``sample_xlsx`` — single ``"Investments"`` sheet, 12 monthly dates, 3 funds.
* ``sample_xlsx_sparse`` — same layout but with one fully-empty column.
"""

from __future__ import annotations

import datetime
import pathlib

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Excel-import workbook builder (shared by both import-format fixtures)
# ---------------------------------------------------------------------------

#: Sheet names in the canonical import-format order.
_V2_SHEET_NAMES: list[str] = [
    "Attributes",
    "Cash Flow In actual",
    "Cash Flow In plan",
    "Cash Flow Out actual",
    "Cash Flow Out plan",
    "NAVs actual",
    "NAVs plan",
    "total return actual",
    "total return plan",
    "interest rates",
]

#: Rate series names in the order they appear in the ``interest rates`` sheet.
_RATE_SERIES_NAMES: list[str] = [
    "risk free rate",
    "benchmark rate",
    "EURIBOR 3M",
    "swap rate 10Y",
]

#: Deterministic daily interest rates (decimal form, 2–5% range) for the fixture.
_RATE_VALUES: list[float] = [
    0.04006,
    0.03985,
    0.04012,
    0.03998,
    0.04021,
    0.03975,
    0.04034,
    0.04011,
    0.03992,
    0.04018,
]

#: 10 consecutive daily dates used as the time-series range in fixtures.
_DATES: list[datetime.datetime] = [datetime.datetime(2024, 1, d) for d in range(1, 11)]

#: Deterministic daily return values for cycling across active investments.
_RETURNS: list[float] = [
    0.0012,
    0.0023,
    -0.0031,
    0.0045,
    -0.0011,
    0.0033,
    -0.0022,
    0.0014,
    0.0025,
    -0.0013,
]


def _investment_names(n: int) -> list[str]:
    """Return ``n`` investment names of the form ``"Investition A"``, etc."""
    return [f"Investition {chr(65 + i)}" for i in range(n)]


def _build_v2_workbook(
    n_investments: int,
    n_active: int,
    n_rate_columns: int = 1,
) -> openpyxl.Workbook:
    """Build a complete Excel import workbook with ``n_investments`` investment columns.

    Produces deterministic, realistic data so that value-level tests are
    possible without depending on external files.  Includes an ``interest rates``
    sheet with ``n_rate_columns`` rate series columns to exercise market reference
    data parsing.

    Args:
        n_investments: Total number of investment columns (active + placeholder).
        n_active: Number of columns that carry real data.  The remaining
            ``n_investments - n_active`` columns are placeholder slots.
        n_rate_columns: Number of rate series columns in the ``interest rates``
            sheet.  Must not exceed ``len(_RATE_SERIES_NAMES)``.

    Returns:
        Populated :class:`openpyxl.Workbook` ready to be saved.
    """
    assert n_active <= n_investments, "n_active cannot exceed n_investments"
    assert n_rate_columns <= len(_RATE_SERIES_NAMES), (
        f"n_rate_columns={n_rate_columns} exceeds available rate series names"
    )
    n_placeholder = n_investments - n_active

    names = _investment_names(n_investments)

    # Types/sub-classes for active investments; placeholder text for the rest
    types = [
        "Aktien",
        "Private Equity",
        "Real Estate",
        "Immobilien",
        "Infrastruktur",
        "Hedgefonds",
        "Rohstoffe",
        "Anleihen",
    ][:n_active] + ["Typ der Investition"] * n_placeholder
    subclasses = [
        "Large Cap Defensiv",
        "Buyout",
        "Core",
        "Wohnen",
        "Energie",
        "Long/Short",
        "Edelmetalle",
        "Staatsanleihen",
    ][:n_active] + ["Klasse der Investition"] * n_placeholder

    # Shared header rows written to every sheet
    header_row_1 = [None, *names]
    header_row_2 = [None, *types]
    header_row_3 = [None, *subclasses]

    wb = openpyxl.Workbook()
    # Remove the default blank sheet; we will create all sheets explicitly
    wb.remove(wb.active)
    for name in _V2_SHEET_NAMES:
        wb.create_sheet(name)

    # ------------------------------------------------------------------
    # Attributes sheet
    # ------------------------------------------------------------------
    ws = wb["Attributes"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)

    attribute_data = [
        (
            "Region",
            ["Europa", "Europa", "Nordamerika", "Asien", "Europa", "USA", "Global", "Europa"][
                :n_active
            ],
        ),
        ("Vintage Year", [2020, 2021, 2019, 2022, 2018, 2020, 2019, 2017][:n_active]),
        ("Währung", ["EUR", "EUR", "USD", "EUR", "USD", "USD", "EUR", "EUR"][:n_active]),
        ("Asset Class", types[:n_active]),
        ("Manager / Fondsname", [f"Manager {chr(65 + i)} GmbH" for i in range(n_active)]),
    ]
    for label, values in attribute_data:
        ws.append([label] + values + [None] * n_placeholder)

    # ------------------------------------------------------------------
    # Cash Flow In actual — positive values for active investments
    # ------------------------------------------------------------------
    ws = wb["Cash Flow In actual"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for i, date in enumerate(_DATES):
        values = [50_000 * (j + 1) * (i + 1) for j in range(n_active)]
        ws.append([date] + values + [None] * n_placeholder)

    # ------------------------------------------------------------------
    # Cash Flow In plan — empty (dates present, all values None)
    # ------------------------------------------------------------------
    ws = wb["Cash Flow In plan"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for date in _DATES:
        ws.append([date] + [None] * n_investments)

    # ------------------------------------------------------------------
    # Cash Flow Out actual — negative initial investment on the first date
    # ------------------------------------------------------------------
    ws = wb["Cash Flow Out actual"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for i, date in enumerate(_DATES):
        if i == 0:
            # Record the initial outflow on day 1; all later rows are None
            values: list[float | None] = [-1_000_000 * (j + 1) for j in range(n_active)]
        else:
            values = [None] * n_active  # type: ignore[assignment]
        ws.append([date] + values + [None] * n_placeholder)

    # ------------------------------------------------------------------
    # Cash Flow Out plan — empty (dates present, all values None)
    # ------------------------------------------------------------------
    ws = wb["Cash Flow Out plan"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for date in _DATES:
        ws.append([date] + [None] * n_investments)

    # ------------------------------------------------------------------
    # NAVs actual — positive absolute amounts, growing over time
    # ------------------------------------------------------------------
    ws = wb["NAVs actual"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for i, date in enumerate(_DATES):
        # Each investment starts at a different base NAV and grows at 1 %/day
        nav_values = [round(1_000_000 * (j + 1) * (1 + 0.01 * i), 2) for j in range(n_active)]
        ws.append([date] + nav_values + [None] * n_placeholder)

    # ------------------------------------------------------------------
    # NAVs plan — empty (dates present, all values None)
    # ------------------------------------------------------------------
    ws = wb["NAVs plan"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for date in _DATES:
        ws.append([date] + [None] * n_investments)

    # ------------------------------------------------------------------
    # total return actual — small decimal returns
    # ------------------------------------------------------------------
    ws = wb["total return actual"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for i, date in enumerate(_DATES):
        ret_values = [round(_RETURNS[i] * (j + 1), 6) for j in range(n_active)]
        ws.append([date] + ret_values + [None] * n_placeholder)

    # ------------------------------------------------------------------
    # total return plan — ENTIRELY EMPTY (dates present, all values None)
    # This sheet is used by test_empty_plan_sheet to verify that the parser
    # returns an all-NaN DataFrame rather than raising.
    # ------------------------------------------------------------------
    ws = wb["total return plan"]
    ws.append(header_row_1)
    ws.append(header_row_2)
    ws.append(header_row_3)
    for date in _DATES:
        ws.append([date] + [None] * n_investments)

    # ------------------------------------------------------------------
    # interest rates — market reference data, independent column namespace.
    # Row 1: column A empty, then rate series names.
    # Rows 2–3: entirely empty (no investment type/sub-class metadata).
    # Rows 4+: dates and decimal annual rate values (2–5% range).
    # ------------------------------------------------------------------
    rate_names = _RATE_SERIES_NAMES[:n_rate_columns]
    ws = wb["interest rates"]
    ws.append([None, *rate_names])  # row 1: rate series names, column A empty
    ws.append([None] * (n_rate_columns + 1))  # row 2: empty
    ws.append([None] * (n_rate_columns + 1))  # row 3: empty
    for i, date in enumerate(_DATES):
        # Cycle through _RATE_VALUES and apply a small offset per series so
        # each column carries distinct but realistic values
        rate_row = [round(_RATE_VALUES[i] + j * 0.005, 6) for j in range(n_rate_columns)]
        ws.append([date, *rate_row])

    return wb


# ---------------------------------------------------------------------------
# Excel-import-format fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_excel_import_workbook(tmp_path: pathlib.Path) -> pathlib.Path:
    """Create an Excel import workbook with 5 investment columns (3 active, 2 placeholders).

    Also contains an ``interest rates`` sheet with 1 rate series column
    (``"risk free rate"``), used to test market reference data parsing.

    Using a column count that differs from the production file is intentional:
    it proves that the parser discovers column counts dynamically rather than
    relying on any hardcoded assumption.

    Args:
        tmp_path: Pytest-provided temporary directory.

    Returns:
        :class:`pathlib.Path` pointing to the written ``.xlsx`` file.
    """
    wb = _build_v2_workbook(n_investments=5, n_active=3, n_rate_columns=1)
    path = tmp_path / "test_v2_5col.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def sample_excel_import_workbook_extended(tmp_path: pathlib.Path) -> pathlib.Path:
    """Create an Excel import workbook with 8 investment columns (6 active, 2 placeholders).

    Also contains an ``interest rates`` sheet with 2 rate series columns
    (``"risk free rate"``, ``"benchmark rate"``), used to test dynamic column
    discovery for market reference sheets.

    Used in ``test_dynamic_column_count``, ``test_column_count_differs_from_fixture``,
    and ``test_interest_rates_dynamic_columns`` to explicitly prove that the parser
    adapts to different column counts without any code changes.

    Args:
        tmp_path: Pytest-provided temporary directory.

    Returns:
        :class:`pathlib.Path` pointing to the written ``.xlsx`` file.
    """
    wb = _build_v2_workbook(n_investments=8, n_active=6, n_rate_columns=2)
    path = tmp_path / "test_v2_8col.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Legacy V1 fixtures — kept for backward-compatibility tests
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_xlsx(tmp_path: pathlib.Path) -> pathlib.Path:
    """[LEGACY V1] Create a single-sheet Excel workbook.

    The workbook contains a single sheet named ``"Investments"`` with:

    * Column 0 (``"Date"``): 12 monthly timestamps from 2020-01-01 to 2020-12-01.
    * Columns 1–3: ``"Fund Alpha"``, ``"Fund Beta"``, ``"Fund Gamma"`` with
      deterministic float return values.

    This fixture is retained for legacy tests that exercise V1-style behaviour
    (e.g. validating that a non-DatetimeIndex raises ``ValidationError``).  It
    is **not** a valid Excel import workbook and cannot be passed to the
    multi-sheet ``load_excel``.

    Args:
        tmp_path: Pytest-provided temporary directory.

    Returns:
        :class:`pathlib.Path` pointing to the written ``.xlsx`` file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Investments"

    headers = ["Date", "Fund Alpha", "Fund Beta", "Fund Gamma"]
    ws.append(headers)

    monthly_returns = [
        (0.0232, -0.0045, 0.0314),
        (0.0376, 0.0112, -0.0189),
        (-0.0120, 0.0251, 0.0087),
        (0.0445, -0.0318, 0.0229),
        (0.0089, 0.0195, -0.0067),
        (-0.0234, 0.0321, 0.0154),
        (0.0312, -0.0187, 0.0276),
        (0.0156, 0.0098, -0.0143),
        (-0.0078, 0.0267, 0.0321),
        (0.0421, -0.0234, 0.0087),
        (0.0198, 0.0145, -0.0256),
        (-0.0089, 0.0312, 0.0198),
    ]
    for month_idx in range(12):
        date = datetime.datetime(2020, month_idx + 1, 1)
        returns = monthly_returns[month_idx]
        ws.append([date, *returns])

    path = tmp_path / "test_investments_legacy.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def sample_xlsx_sparse(tmp_path: pathlib.Path) -> pathlib.Path:
    """[LEGACY V1] Create a single-sheet workbook with one all-empty column.

    Args:
        tmp_path: Pytest-provided temporary directory.

    Returns:
        :class:`pathlib.Path` to the written workbook.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Investments"

    ws.append(["Date", "Fund Alpha", "Fund Empty", "Fund Beta"])
    for month_idx in range(6):
        date = datetime.datetime(2021, month_idx + 1, 1)
        ws.append(
            [
                date,
                round(0.01 * (month_idx + 1), 4),
                None,  # Fund Empty — entirely absent
                round(-0.005 * (month_idx + 1), 4),
            ]
        )

    path = tmp_path / "test_sparse_legacy.xlsx"
    wb.save(path)
    return path

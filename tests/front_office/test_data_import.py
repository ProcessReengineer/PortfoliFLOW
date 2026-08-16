# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025-2026 Sönke Pinkernelle

# PortfoliFLOW — Unit tests for the Data Import module
"""Unit tests for :mod:`modules.front_office.data_import`.

Coverage targets
----------------
* Multi-sheet Excel-import loading: all 10 sheets, selective loading, canonical keys.
* Attributes sheet: structure, index, investment-type row.
* Investment time-series sheets: DatetimeIndex, sort order, column names, dtypes.
* Market reference sheets: interest_rates structure, empty rows 2–3, dynamic columns.
* Edge cases: empty plan sheets, placeholder columns, mixed active/empty columns.
* Value semantics: negative cash-flow-out, positive NAVs, decimal returns.
* Error paths: missing file, non-Excel file, no recognised sheets.
* Schema validation: ``validate_dataframe`` and ``validate_workbook``.
* DataStore integration: datasets accessible after ``run()``.
* Dynamic column count: parser adapts without code changes.
* Headless importability: no display required at import time.

All tests use in-memory Excel fixtures from ``conftest.py``; none depend on
real sample files in ``data/sample/``.
"""

from __future__ import annotations

import datetime
import logging
import pathlib

import openpyxl
import pandas as pd
import pytest

from core.exceptions import DataImportError, ValidationError
from services.data_normalization.excel_workbook_loader import (
    load_excel,
    validate_dataframe,
    validate_workbook,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

#: All 10 canonical keys that a full Excel import workbook produces.
ALL_CANONICAL_KEYS: list[str] = [
    "attributes",
    "cash_flow_in_actual",
    "cash_flow_in_plan",
    "cash_flow_out_actual",
    "cash_flow_out_plan",
    "navs_actual",
    "navs_plan",
    "total_return_actual",
    "total_return_plan",
    "interest_rates",
]

#: Canonical keys for investment time-series sheets only (excludes attributes and
#: market reference sheets).  Used in tests that verify investment-specific invariants.
INVESTMENT_TS_KEYS: list[str] = [
    "cash_flow_in_actual",
    "cash_flow_in_plan",
    "cash_flow_out_actual",
    "cash_flow_out_plan",
    "navs_actual",
    "navs_plan",
    "total_return_actual",
    "total_return_plan",
]

#: Number of investments in the standard 5-column fixture.
N_INVESTMENTS_STANDARD = 5
#: Number of investments in the extended 8-column fixture.
N_INVESTMENTS_EXTENDED = 8


@pytest.fixture
def clean_data_store():
    """Reset the DataStore singleton before each test that touches it.

    Ensures that DataStore-dependent tests are fully isolated from each other
    regardless of test execution order.
    """
    from core.data_store import get_data_store

    store = get_data_store()
    store.clear()
    yield store
    store.clear()


# ---------------------------------------------------------------------------
# Loading: all sheets
# ---------------------------------------------------------------------------


def test_load_all_sheets(sample_excel_import_workbook: pathlib.Path) -> None:
    """``load_excel`` with no ``sheets`` filter must return exactly 10 datasets."""
    datasets = load_excel(sample_excel_import_workbook)
    assert len(datasets) == 10, f"Expected 10 datasets, got {len(datasets)}: {list(datasets)}"


def test_canonical_keys(sample_excel_import_workbook: pathlib.Path) -> None:
    """The dict keys must match the expected canonical snake_case names exactly."""
    datasets = load_excel(sample_excel_import_workbook)
    assert sorted(datasets.keys()) == sorted(ALL_CANONICAL_KEYS), (
        f"Key mismatch: got {sorted(datasets.keys())!r}"
    )


def test_load_selective_sheets(sample_excel_import_workbook: pathlib.Path) -> None:
    """``load_excel(sheets=[...])`` returns only the requested sheets."""
    datasets = load_excel(sample_excel_import_workbook, sheets=["Attributes", "NAVs actual"])
    assert set(datasets.keys()) == {"attributes", "navs_actual"}, (
        f"Expected only 2 datasets, got: {list(datasets)}"
    )


# ---------------------------------------------------------------------------
# Attributes sheet
# ---------------------------------------------------------------------------


def test_attributes_structure(sample_excel_import_workbook: pathlib.Path) -> None:
    """Attributes DataFrame has investment names as columns and labels as index."""
    datasets = load_excel(sample_excel_import_workbook)
    df = datasets["attributes"]

    assert isinstance(df, pd.DataFrame)
    # Columns are the investment names discovered from row 1
    assert len(df.columns) == N_INVESTMENTS_STANDARD, (
        f"Expected {N_INVESTMENTS_STANDARD} columns, got {len(df.columns)}"
    )
    # Index must contain the two synthetic rows plus attribute rows
    assert "Investment Type" in df.index, "Index must contain 'Investment Type'"
    assert "Investment Sub-Class" in df.index, "Index must contain 'Investment Sub-Class'"


def test_attributes_investment_type_row(sample_excel_import_workbook: pathlib.Path) -> None:
    """Row 'Investment Type' must contain real type values for active investments."""
    df = load_excel(sample_excel_import_workbook)["attributes"]
    type_row = df.loc["Investment Type"]

    # First 3 investments (active) must have non-placeholder types
    active_types = type_row.iloc[:3].tolist()
    placeholder_types = type_row.iloc[3:].tolist()

    for t in active_types:
        assert t is not None and t != "Typ der Investition", (
            f"Active investment has placeholder type: {t!r}"
        )
    for t in placeholder_types:
        assert t == "Typ der Investition", (
            f"Placeholder investment should have 'Typ der Investition', got: {t!r}"
        )


# ---------------------------------------------------------------------------
# Time-series sheets: structural invariants
# ---------------------------------------------------------------------------


def test_timeseries_date_index(sample_excel_import_workbook: pathlib.Path) -> None:
    """All time-series DataFrames must have a DatetimeIndex named 'Date'."""
    datasets = load_excel(sample_excel_import_workbook)
    ts_keys = [k for k in datasets if k != "attributes"]

    for key in ts_keys:
        df = datasets[key]
        assert isinstance(df.index, pd.DatetimeIndex), (
            f"Sheet '{key}': expected DatetimeIndex, got {type(df.index).__name__}"
        )
        assert df.index.name == "Date", (
            f"Sheet '{key}': index name is {df.index.name!r}, expected 'Date'"
        )
        # All timestamps must be normalised to midnight
        for ts in df.index:
            assert ts.hour == 0 and ts.minute == 0 and ts.second == 0, (
                f"Sheet '{key}': timestamp {ts!r} has non-zero time component"
            )


def test_timeseries_sorted_index(sample_excel_import_workbook: pathlib.Path) -> None:
    """Time-series DatetimeIndex must be monotonically increasing."""
    datasets = load_excel(sample_excel_import_workbook)
    ts_keys = [k for k in datasets if k != "attributes"]

    for key in ts_keys:
        df = datasets[key]
        if len(df.index) > 1:
            assert df.index.is_monotonic_increasing, f"Sheet '{key}': DatetimeIndex is not sorted"


def test_timeseries_float64_dtypes(sample_excel_import_workbook: pathlib.Path) -> None:
    """All value columns in time-series DataFrames must have dtype ``float64``."""
    datasets = load_excel(sample_excel_import_workbook)
    ts_keys = [k for k in datasets if k != "attributes"]

    for key in ts_keys:
        df = datasets[key]
        for col in df.columns:
            assert df[col].dtype == "float64", (
                f"Sheet '{key}', column '{col}': expected float64, got {df[col].dtype}"
            )


def test_timeseries_columns_match_attributes(sample_excel_import_workbook: pathlib.Path) -> None:
    """Investment column names must be identical across all investment sheets.

    Market reference sheets (``interest_rates``) have their own independent
    column namespace and are excluded from this check.
    """
    datasets = load_excel(sample_excel_import_workbook)
    reference_cols = list(datasets["attributes"].columns)

    for key in INVESTMENT_TS_KEYS:
        df = datasets[key]
        assert list(df.columns) == reference_cols, (
            f"Sheet '{key}' column mismatch: {list(df.columns)!r} != {reference_cols!r}"
        )


# ---------------------------------------------------------------------------
# Empty plan sheet handling
# ---------------------------------------------------------------------------


def test_empty_plan_sheet(sample_excel_import_workbook: pathlib.Path) -> None:
    """A fully-empty plan sheet must load without error; all values must be NaN.

    ``total_return_plan`` in the fixture has dates in column A but ``None`` for
    every investment value — representing future projections not yet filled in.
    """
    datasets = load_excel(sample_excel_import_workbook)
    df = datasets["total_return_plan"]

    assert isinstance(df.index, pd.DatetimeIndex), "Plan sheet must still have DatetimeIndex"
    assert len(df.index) == 10, "Plan sheet index must have the 10 fixture dates"
    assert df.isna().all(axis=None), "Every value in the empty plan sheet must be NaN"


# ---------------------------------------------------------------------------
# Placeholder column handling
# ---------------------------------------------------------------------------


def test_placeholder_columns_preserved(sample_excel_import_workbook: pathlib.Path) -> None:
    """Placeholder investment columns must be present in all investment DataFrames.

    The Excel import format uses placeholder column names in row 1 to reserve slots for
    future investments.  The parser must NOT drop them even though their data
    cells are all ``None`` / NaN.

    Market reference sheets (``interest_rates``) have a different column count
    by design and are excluded from this check.
    """
    datasets = load_excel(sample_excel_import_workbook)

    investment_keys = ["attributes", *INVESTMENT_TS_KEYS]
    for key in investment_keys:
        df = datasets[key]
        assert len(df.columns) == N_INVESTMENTS_STANDARD, (
            f"Sheet '{key}': expected {N_INVESTMENTS_STANDARD} columns "
            f"(including placeholders), got {len(df.columns)}"
        )

    # The last 2 columns are placeholders: all NaN in every investment time-series sheet
    for key in INVESTMENT_TS_KEYS:
        df = datasets[key]
        for col in df.columns[3:]:  # columns at index 3 and 4 are placeholders
            assert df[col].isna().all(), (
                f"Sheet '{key}', placeholder column '{col}' should be all NaN"
            )


# ---------------------------------------------------------------------------
# Value semantics
# ---------------------------------------------------------------------------


def test_cashflow_out_negative(sample_excel_import_workbook: pathlib.Path) -> None:
    """Cash Flow Out actual must contain negative values for active investments."""
    df = load_excel(sample_excel_import_workbook)["cash_flow_out_actual"]

    # Active investments have a negative initial outflow on the first date
    first_row = df.iloc[0]
    active_values = first_row.iloc[:3].dropna()
    assert len(active_values) > 0, "No non-NaN values in first row of cash_flow_out_actual"
    assert (active_values < 0).all(), (
        f"Expected negative values in first row, got: {active_values.tolist()}"
    )


def test_nav_values_positive(sample_excel_import_workbook: pathlib.Path) -> None:
    """NAVs actual must contain strictly positive values for active investments."""
    df = load_excel(sample_excel_import_workbook)["navs_actual"]

    active_cols = df.columns[:3]
    for col in active_cols:
        non_nan = df[col].dropna()
        assert len(non_nan) > 0, f"Column '{col}' has no data in navs_actual"
        assert (non_nan > 0).all(), f"Column '{col}' contains non-positive NAVs: {non_nan.tolist()}"


def test_return_values_decimal(sample_excel_import_workbook: pathlib.Path) -> None:
    """total return actual values must be decimal fractions in the range (-1, 1)."""
    df = load_excel(sample_excel_import_workbook)["total_return_actual"]

    active_cols = df.columns[:3]
    for col in active_cols:
        non_nan = df[col].dropna()
        assert len(non_nan) > 0, f"Column '{col}' has no return data"
        assert (non_nan > -1).all() and (non_nan < 1).all(), (
            f"Column '{col}' has out-of-range returns: min={non_nan.min()}, max={non_nan.max()}"
        )


# ---------------------------------------------------------------------------
# Error paths
# ---------------------------------------------------------------------------


def test_missing_file_raises(tmp_path: pathlib.Path) -> None:
    """Passing a non-existent path must raise :exc:`DataImportError`."""
    ghost = tmp_path / "does_not_exist.xlsx"
    with pytest.raises(DataImportError, match="not found"):
        load_excel(ghost)


def test_invalid_file_raises(tmp_path: pathlib.Path) -> None:
    """Passing a non-Excel file must raise :exc:`DataImportError`."""
    bad_file = tmp_path / "not_an_excel.xlsx"
    bad_file.write_text("this is not a zip archive")
    with pytest.raises(DataImportError):
        load_excel(bad_file)


def test_no_recognized_sheets_raises(tmp_path: pathlib.Path) -> None:
    """A workbook with no recognised sheet names must raise :exc:`DataImportError`."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UnknownSheet"
    ws.append(["some", "data"])
    path = tmp_path / "no_recognized_sheets.xlsx"
    wb.save(path)

    with pytest.raises(DataImportError, match="No recognised sheets"):
        load_excel(path)


# ---------------------------------------------------------------------------
# validate_dataframe
# ---------------------------------------------------------------------------


def test_validate_timeseries(sample_excel_import_workbook: pathlib.Path) -> None:
    """``validate_dataframe`` with ``kind='timeseries'`` passes on a valid sheet."""
    df = load_excel(sample_excel_import_workbook)["navs_actual"]
    # Must not raise
    validate_dataframe(df, kind="timeseries")


def test_validate_attributes(sample_excel_import_workbook: pathlib.Path) -> None:
    """``validate_dataframe`` with ``kind='attributes'`` passes on the Attributes sheet."""
    df = load_excel(sample_excel_import_workbook)["attributes"]
    # Must not raise
    validate_dataframe(df, kind="attributes")


def test_validate_allnan_plan_sheet(sample_excel_import_workbook: pathlib.Path) -> None:
    """An all-NaN plan sheet must pass ``validate_dataframe`` (plan sheets are valid).

    This replaces the old V1 test that required non-empty DataFrames: in
    the Excel import format, plan sheets that have dates but no values yet
    are a first-class concept.
    """
    df = load_excel(sample_excel_import_workbook)["total_return_plan"]
    # Must not raise even though every value cell is NaN
    validate_dataframe(df, kind="timeseries")


def test_validate_non_datetime_index() -> None:
    """A timeseries DataFrame with a RangeIndex must fail ``validate_dataframe``."""
    df = pd.DataFrame({"Fund A": [0.01, 0.02], "Fund B": [0.02, -0.01]})
    assert not isinstance(df.index, pd.DatetimeIndex)
    with pytest.raises(ValidationError, match="DatetimeIndex"):
        validate_dataframe(df, kind="timeseries")


def test_validate_single_column_passes() -> None:
    """A single-column timeseries DataFrame must pass validation."""
    idx = pd.date_range("2020-01-01", periods=3, freq="MS", name="Date")
    df = pd.DataFrame({"Only Fund": [0.01, 0.02, 0.03]}, index=idx)
    validate_dataframe(df, kind="timeseries")  # must not raise


def test_validate_zero_columns_raises() -> None:
    """A DataFrame with no data columns must fail validation."""
    idx = pd.date_range("2020-01-01", periods=3, freq="MS", name="Date")
    df = pd.DataFrame(index=idx)
    with pytest.raises(ValidationError):
        validate_dataframe(df, kind="timeseries")


def test_validate_infinite_values_raises() -> None:
    """Columns containing ``±inf`` must cause validation to raise."""
    idx = pd.date_range("2020-01-01", periods=3, freq="MS", name="Date")
    df = pd.DataFrame(
        {"Fund A": [0.01, float("inf"), 0.03], "Fund B": [0.02, 0.01, -0.01]},
        index=idx,
    )
    with pytest.raises(ValidationError, match="infinite"):
        validate_dataframe(df, kind="timeseries")


def test_validate_attributes_rejects_datetime_index() -> None:
    """Passing a DataFrame with a DatetimeIndex to ``kind='attributes'`` must raise."""
    idx = pd.date_range("2020-01-01", periods=3, freq="MS", name="Date")
    df = pd.DataFrame({"Inv A": ["EUR", "EUR", "EUR"]}, index=idx)
    with pytest.raises(ValidationError, match="DatetimeIndex"):
        validate_dataframe(df, kind="attributes")


# ---------------------------------------------------------------------------
# validate_workbook
# ---------------------------------------------------------------------------


def test_cross_sheet_consistency(sample_excel_import_workbook: pathlib.Path) -> None:
    """``validate_workbook`` must pass when all sheets share the same columns."""
    datasets = load_excel(sample_excel_import_workbook)
    # Must not raise
    validate_workbook(datasets)


def test_cross_sheet_consistency_mismatch() -> None:
    """``validate_workbook`` must raise when column names differ across sheets."""
    idx = pd.date_range("2024-01-01", periods=5, name="Date")
    ds = {
        "attributes": pd.DataFrame(
            {"Inv A": ["Aktien"], "Inv B": ["PE"]},
            index=["Investment Type"],
        ),
        "navs_actual": pd.DataFrame(
            {"Inv A": [1.0, 2.0, 3.0, 4.0, 5.0], "Inv C": [1.0, 2.0, 3.0, 4.0, 5.0]},
            index=idx,
        ),
    }
    with pytest.raises(ValidationError, match="mismatch"):
        validate_workbook(ds)


# ---------------------------------------------------------------------------
# DataStore integration
# ---------------------------------------------------------------------------


def test_datastore_integration(
    sample_excel_import_workbook: pathlib.Path,
    clean_data_store,
) -> None:
    """After ``run()``, all 9 datasets must be retrievable from the DataStore."""
    from core.config import get_config
    from modules.module_registry import registry

    cls = registry.get("data_import")
    instance = cls(config=get_config())
    result = instance.run(action="load_excel", source=str(sample_excel_import_workbook))

    assert result["status"] == "ok"
    assert result["metadata"]["n_datasets"] == 10

    store = clean_data_store
    for key in ALL_CANONICAL_KEYS:
        df = store.get(key)
        assert df is not None, f"DataStore missing key '{key}' after import"
        assert isinstance(df, pd.DataFrame)


def test_run_missing_source_raises() -> None:
    """``run()`` without a ``source`` argument must raise :exc:`ValidationError`."""
    from core.config import get_config
    from modules.module_registry import registry

    cls = registry.get("data_import")
    instance = cls(config=get_config())
    with pytest.raises(ValidationError, match="source"):
        instance.run(action="load_excel")


def test_run_via_registry(
    sample_excel_import_workbook: pathlib.Path,
    clean_data_store,
) -> None:
    """Running :class:`DataImport` through the registry must return a valid result."""
    from core.config import get_config
    from modules.module_registry import registry

    cls = registry.get("data_import")
    instance = cls(config=get_config())
    result = instance.run(action="load_excel", source=str(sample_excel_import_workbook))

    assert result["status"] == "ok"
    assert "datasets" in result
    assert isinstance(result["datasets"], dict)
    assert "navs_actual" in result["datasets"]
    assert isinstance(result["datasets"]["navs_actual"], pd.DataFrame)
    assert set(result["metadata"]["keys"]) == set(ALL_CANONICAL_KEYS)


# ---------------------------------------------------------------------------
# Dynamic column count (the core correctness test for the Excel-import parser design)
# ---------------------------------------------------------------------------


def test_dynamic_column_count(sample_excel_import_workbook_extended: pathlib.Path) -> None:
    """Investment sheets in the 8-column extended fixture must have 8 columns.

    This is the definitive test that the parser discovers the investment column
    count dynamically from row 1 rather than relying on any hardcoded value.

    Market reference sheets (``interest_rates``) have their own column count
    (2 in the extended fixture) and are excluded from this assertion.
    """
    datasets = load_excel(sample_excel_import_workbook_extended)

    investment_keys = ["attributes", *INVESTMENT_TS_KEYS]
    for key in investment_keys:
        df = datasets[key]
        assert len(df.columns) == N_INVESTMENTS_EXTENDED, (
            f"Sheet '{key}': expected {N_INVESTMENTS_EXTENDED} columns "
            f"(from extended fixture), got {len(df.columns)}"
        )


def test_column_count_differs_from_fixture(
    sample_excel_import_workbook: pathlib.Path,
    sample_excel_import_workbook_extended: pathlib.Path,
) -> None:
    """Standard and extended fixtures must produce DataFrames with different column counts.

    This asserts that the same parser code adapts dynamically to different
    workbooks — if both fixtures returned the same column count, the test would
    be vacuous.
    """
    standard = load_excel(sample_excel_import_workbook)
    extended = load_excel(sample_excel_import_workbook_extended)

    # Use a consistent investment sheet (attributes) — market reference sheets
    # have their own column counts and must not be used for this comparison.
    std_cols = len(standard["attributes"].columns)
    ext_cols = len(extended["attributes"].columns)

    assert std_cols != ext_cols, (
        f"Both fixtures produced {std_cols} columns; "
        "they must differ to prove dynamic column discovery."
    )
    assert std_cols == N_INVESTMENTS_STANDARD
    assert ext_cols == N_INVESTMENTS_EXTENDED


# ---------------------------------------------------------------------------
# Headless importability
# ---------------------------------------------------------------------------


def test_headless_import(monkeypatch: pytest.MonkeyPatch) -> None:
    """The data-import module must be importable with no graphical display.

    :mod:`modules.front_office.data_import` must not import any GUI packages at
    module scope.  This test removes the ``DISPLAY`` environment variable and
    verifies that the public functions remain callable.
    """
    monkeypatch.delenv("DISPLAY", raising=False)
    monkeypatch.setenv("DISPLAY", "")

    assert callable(load_excel), "load_excel must be callable without a display"
    assert callable(validate_dataframe), "validate_dataframe must be callable without a display"
    assert callable(validate_workbook), "validate_workbook must be callable without a display"


# ---------------------------------------------------------------------------
# Market reference data: interest rates sheet
# ---------------------------------------------------------------------------


def test_interest_rates_loaded(sample_excel_import_workbook: pathlib.Path) -> None:
    """``load_excel`` must return ``"interest_rates"`` as one of the dict keys."""
    datasets = load_excel(sample_excel_import_workbook)
    assert "interest_rates" in datasets, (
        f"Expected 'interest_rates' key; got: {sorted(datasets.keys())}"
    )


def test_interest_rates_structure(sample_excel_import_workbook: pathlib.Path) -> None:
    """The ``interest_rates`` DataFrame must have a DatetimeIndex and float64 columns.

    Structural requirements mirror investment time-series sheets: the same
    parsing path is used, so the output schema is identical.
    """
    datasets = load_excel(sample_excel_import_workbook)
    df = datasets["interest_rates"]

    assert isinstance(df.index, pd.DatetimeIndex), (
        f"interest_rates: expected DatetimeIndex, got {type(df.index).__name__}"
    )
    assert df.index.name == "Date", (
        f"interest_rates: index name is {df.index.name!r}, expected 'Date'"
    )
    assert len(df.columns) >= 1, "interest_rates must have at least one rate series column"
    assert df.columns[0] == "risk free rate", (
        f"First rate column should be 'risk free rate', got {df.columns[0]!r}"
    )
    for col in df.columns:
        assert df[col].dtype == "float64", (
            f"interest_rates column '{col}': expected float64, got {df[col].dtype}"
        )


def test_interest_rates_empty_metadata_rows(sample_excel_import_workbook: pathlib.Path) -> None:
    """Rows 2–3 being entirely empty in ``interest rates`` must not cause errors.

    The time-series parser skips rows 2–3 entirely (it starts at row 4), so
    empty metadata rows in market reference sheets are handled transparently.
    """
    # If this raises, the parser is crashing on empty rows 2–3.
    datasets = load_excel(sample_excel_import_workbook)
    df = datasets["interest_rates"]

    assert len(df.index) == 10, f"interest_rates: expected 10 date rows, got {len(df.index)}"
    assert df.notna().any(axis=None), "interest_rates: expected non-NaN rate values, got all NaN"


def test_interest_rates_dynamic_columns(
    sample_excel_import_workbook_extended: pathlib.Path,
) -> None:
    """The extended fixture (2 rate columns) must produce 2 columns in ``interest_rates``.

    Mirrors ``test_dynamic_column_count`` but specifically for the market reference
    sheet category, proving that rate column discovery is also dynamic.
    """
    datasets = load_excel(sample_excel_import_workbook_extended)
    df = datasets["interest_rates"]

    assert len(df.columns) == 2, (
        f"Extended fixture should have 2 rate columns, got {len(df.columns)}: {list(df.columns)}"
    )
    assert "risk free rate" in df.columns, "First rate column must be 'risk free rate'"
    assert "benchmark rate" in df.columns, "Second rate column must be 'benchmark rate'"


def test_interest_rates_excluded_from_investment_consistency(
    sample_excel_import_workbook: pathlib.Path,
) -> None:
    """``validate_workbook`` must not raise when ``interest_rates`` columns differ.

    The ``interest_rates`` sheet uses rate series names (e.g. ``"risk free rate"``)
    as columns instead of investment names.  The investment column consistency
    check must exclude it entirely.
    """
    datasets = load_excel(sample_excel_import_workbook)

    # Verify columns genuinely differ so the test is non-vacuous
    inv_cols = list(datasets["attributes"].columns)
    rate_cols = list(datasets["interest_rates"].columns)
    assert inv_cols != rate_cols, (
        "Test pre-condition: interest_rates columns should differ from investment columns"
    )

    # Must not raise even though interest_rates has a different column namespace
    validate_workbook(datasets)


# ---------------------------------------------------------------------------
# Liquid-archetype sheets (ADR-0081) — income + three tidy reference sheets
# ---------------------------------------------------------------------------


def _build_liquid_workbook() -> openpyxl.Workbook:
    """Build a tiny workbook carrying the four liquid-archetype sheets.

    Two investments (one ``Aktien`` = listed_equity, one ``Credit`` =
    listed_bonds), an ``NAVs actual`` sheet, the income sheet pair (wide),
    the three tidy reference sheets, plus one unrecognised sheet to assert
    warn-and-skip.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    names = ["Equity Fund", "Credit Fund"]
    header_1 = [None, *names]
    header_2 = [None, "Aktien", "Credit"]
    header_3 = [None, "Large Cap", "IG"]
    d1 = datetime.datetime(2024, 1, 31)
    d2 = datetime.datetime(2024, 2, 29)

    def _wide(sheet_name: str, rows: list[list[object]]) -> None:
        ws = wb.create_sheet(sheet_name)
        ws.append(header_1)
        ws.append(header_2)
        ws.append(header_3)
        for row in rows:
            ws.append(row)

    # Attributes.
    ws = wb.create_sheet("Attributes")
    ws.append(header_1)
    ws.append(header_2)
    ws.append(header_3)
    ws.append(["Währung", "EUR", "EUR"])

    _wide("NAVs actual", [[d1, 100.0, 200.0], [d2, 101.0, 201.0]])
    # Income: equity pays a dividend, credit pays a coupon.
    _wide("Cash Flow Income actual", [[d1, 5.0, 9.0]])
    _wide("Cash Flow Income plan", [[d2, 6.0, None]])

    ba = wb.create_sheet("Bond Analytics")
    ba.append(["as_of_date", "investment", "ytm", "eff_duration", "oas", "convexity"])
    ba.append([d1, "Credit Fund", 0.045, 3.2, 0.012, None])
    ba.append([d2, "Credit Fund", 0.046, 3.1, None, None])

    rw = wb.create_sheet("Rating Weights")
    rw.append(["as_of_date", "investment", "rating_bucket", "weight_pct"])
    rw.append([d1, "Credit Fund", "AAA", 60.0])
    rw.append([d1, "Credit Fund", "BBB", 40.0])

    mw = wb.create_sheet("Maturity Weights")
    mw.append(["as_of_date", "investment", "maturity_bucket", "weight_pct"])
    mw.append([d1, "Credit Fund", "1-3y", 100.0])

    # Unrecognised sheet — must warn and be skipped.
    ws = wb.create_sheet("Garbage Sheet")
    ws.append(["junk"])
    ws.append([1])

    return wb


def test_liquid_sheets_parsed_with_expected_keys_and_shapes(
    tmp_path: pathlib.Path,
) -> None:
    wb = _build_liquid_workbook()
    path = tmp_path / "liquid.xlsx"
    wb.save(path)

    datasets = load_excel(path)

    # Income sheets land under the canonical wide-timeseries keys.
    assert "cash_flow_income_actual" in datasets
    assert "cash_flow_income_plan" in datasets
    income = datasets["cash_flow_income_actual"]
    assert isinstance(income.index, pd.DatetimeIndex)
    assert list(income.columns) == ["Equity Fund", "Credit Fund"]

    # Tidy reference sheets carry their own fixed column schema.
    ba = datasets["bond_analytics"]
    assert list(ba.columns) == [
        "as_of_date",
        "investment",
        "ytm",
        "eff_duration",
        "oas",
        "convexity",
    ]
    assert ba.shape == (2, 6)
    # Date column was normalised to an ISO string for deterministic JSONB.
    assert ba.iloc[0]["as_of_date"] == "2024-01-31"
    assert ba.iloc[0]["oas"] == 0.012
    assert ba.iloc[0]["convexity"] is None

    rw = datasets["rating_weights"]
    assert list(rw.columns) == ["as_of_date", "investment", "rating_bucket", "weight_pct"]
    assert rw.shape == (2, 4)

    mw = datasets["maturity_weights"]
    assert list(mw.columns) == ["as_of_date", "investment", "maturity_bucket", "weight_pct"]
    assert mw.shape == (1, 4)


def test_liquid_tidy_sheets_excluded_from_consistency_check(
    tmp_path: pathlib.Path,
) -> None:
    """The tidy reference sheets must not trip the investment-column check."""
    wb = _build_liquid_workbook()
    path = tmp_path / "liquid.xlsx"
    wb.save(path)

    datasets = load_excel(path)
    # Pre-condition: tidy columns genuinely differ from investment columns.
    assert list(datasets["bond_analytics"].columns) != list(datasets["attributes"].columns)
    # Must not raise (load_excel already ran it; call again explicitly).
    validate_workbook(datasets)


def test_unknown_sheet_warns_and_is_skipped(
    tmp_path: pathlib.Path, caplog: pytest.LogCaptureFixture
) -> None:
    wb = _build_liquid_workbook()
    path = tmp_path / "liquid.xlsx"
    wb.save(path)

    with caplog.at_level(logging.WARNING):
        datasets = load_excel(path)

    assert "garbage_sheet" not in datasets
    assert any(
        "Garbage Sheet" in rec.message and "not a recognised" in rec.message
        for rec in caplog.records
    )

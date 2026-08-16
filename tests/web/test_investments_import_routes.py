# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (c) 2025-2026 Sönke Pinkernelle

"""HTTP-level tests for ``POST /api/data-uploads/{id}/import-as-investments``.

Live-DB tests against the compose Postgres. The fixtures seed a
sentinel tenant (idempotent), create a user, create an
``unclassified`` asset class via the bootstrap helper, and upload an
in-process Excel import workbook. Each test then drives the import-as-
investments endpoint.

Coverage:

* CSRF-token enforcement (POST without token → 403).
* ``dry_run=true`` returns counts but does not write.
* ``dry_run=false`` writes the normalised investment rows.
* 404 when the upload id is unknown / cross-tenant.
* 400 when the JSONB snapshot is structurally invalid
  (``ImportFormatError`` translated by the route handler).
"""

from __future__ import annotations

import datetime
import io
import os
import pathlib
import re
from collections.abc import AsyncGenerator
from uuid import UUID, uuid4

import openpyxl
import pytest
import pytest_asyncio
from dotenv import load_dotenv
from httpx import ASGITransport, AsyncClient
from sqlalchemy import NullPool, text
from sqlalchemy.ext.asyncio import AsyncEngine, create_async_engine

from cli.bootstrap import install_unclassified_asset_class
from core.repositories import AssetClassRepository, tenant_context
from core.tenant_constants import SENTINEL_TENANT_ID
from services.password_hashing import hash_password
from web.main import create_app
from web.settings import WebSettings

load_dotenv(pathlib.Path(__file__).resolve().parents[2] / ".env")

DATABASE_URL = os.getenv("DATABASE_URL")
DATABASE_URL_SUPERUSER = os.getenv("DATABASE_URL_SUPERUSER")


def _require_db() -> None:
    if not DATABASE_URL or not DATABASE_URL_SUPERUSER:
        pytest.skip(
            "DATABASE_URL and DATABASE_URL_SUPERUSER must be set; "
            "skipping live-DB import-as-investments tests.",
            allow_module_level=False,
        )


# ---------------------------------------------------------------------------
# Fixtures (mirrored from test_data_import_routes.py)
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def fresh_superuser_engine() -> AsyncGenerator[AsyncEngine, None]:
    _require_db()
    engine = create_async_engine(DATABASE_URL_SUPERUSER, future=True, poolclass=NullPool)
    try:
        yield engine
    finally:
        await engine.dispose()


@pytest_asyncio.fixture
async def fresh_app_engine() -> AsyncGenerator[AsyncEngine, None]:
    _require_db()
    engine = create_async_engine(DATABASE_URL, future=True, poolclass=NullPool)
    try:
        yield engine
    finally:
        await engine.dispose()


@pytest_asyncio.fixture
async def reset_schema(
    fresh_superuser_engine: AsyncEngine,
) -> AsyncGenerator[None, None]:
    truncate_sql = text(
        "TRUNCATE TABLE investment_region_weights, "
        "region_country_memberships, regions, "
        "investment_country_weights, "
        "investment_sector_weights, sectors, "
        "investment_cashflows, investment_navs, investments, "
        "saa_correlations, saa_asset_class_inputs, saa_configurations, "
        "asset_classes, "
        "data_upload_sheets, data_uploads, "
        "login_audit, sessions, audit_log, "
        "data_store_entries, users, tenants "
        "RESTART IDENTITY CASCADE"
    )
    async with fresh_superuser_engine.begin() as conn:
        await conn.execute(truncate_sql)
    try:
        yield
    finally:
        async with fresh_superuser_engine.begin() as conn:
            await conn.execute(truncate_sql)


@pytest_asyncio.fixture
async def seeded_user_and_classes(
    fresh_superuser_engine: AsyncEngine,
    fresh_app_engine: AsyncEngine,
    reset_schema: None,
) -> tuple[UUID, str, str]:
    plaintext = "correct-horse-battery-staple"
    user_id = uuid4()
    email = "uploader@example.com"
    async with fresh_superuser_engine.begin() as conn:
        await conn.execute(
            text(
                "INSERT INTO tenants (id, name, subdomain) VALUES (:id, :name, 'minathena-capital') "
                "ON CONFLICT (id) DO NOTHING"
            ),
            {"id": str(SENTINEL_TENANT_ID), "name": "Sentinel Tenant"},
        )
        await conn.execute(
            text(
                """
                INSERT INTO users
                    (id, tenant_id, email, password_hash,
                     roles, is_active)
                VALUES
                    (:id, :tid, :email, :hash, ARRAY['owner']::text[], TRUE)
                """
            ),
            {
                "id": str(user_id),
                "tid": str(SENTINEL_TENANT_ID),
                "email": email,
                "hash": hash_password(plaintext),
            },
        )
    # Install the bootstrap "unclassified" asset class plus a couple
    # of resolvable codes so the importer has somewhere to land.
    async with tenant_context(fresh_app_engine, SENTINEL_TENANT_ID, user_id=user_id) as session:
        repo = AssetClassRepository(session)
        await install_unclassified_asset_class(repo)
        await repo.create(code="private_equity", display_name="Private Equity")
        await repo.create(code="listed_equity", display_name="Listed Equity")
    return user_id, email, plaintext


@pytest_asyncio.fixture
async def web_client(
    seeded_user_and_classes: tuple[UUID, str, str],
) -> AsyncGenerator[AsyncClient, None]:
    settings = WebSettings(
        web_host="127.0.0.1",
        web_port=8000,
        session_cookie_name="portfoliflow_session",
        csrf_cookie_name="portfoliflow_csrf_pre_session",
        database_url=DATABASE_URL,
        database_url_superuser=DATABASE_URL_SUPERUSER,
        session_cookie_secure=False,
    )
    app = create_app(settings)
    transport = ASGITransport(app=app)
    async with (
        AsyncClient(transport=transport, base_url="http://testserver") as client,
        app.router.lifespan_context(app),
    ):
        yield client


async def _login_and_get_csrf(client: AsyncClient, email: str, password: str) -> str:
    """Drive ``GET /login`` + ``POST /login``, return session CSRF.

    Scrapes the token from the Admin area page, which now embeds
    the upload form (relocated from Front Office in the 6F-3
    mid-polish; the standalone ``/data-import`` route was sunset
    in sub-stream 6F-5).
    """
    get_response = await client.get("/login")
    csrf = get_response.cookies.get("portfoliflow_csrf_pre_session")
    assert csrf is not None
    await client.post(
        "/login",
        data={"email": email, "password": password, "csrf_token": csrf},
        follow_redirects=False,
    )
    page = await client.get("/admin", follow_redirects=False)
    assert page.status_code == 200
    match = re.search(r'name="csrf_token"\s+value="([^"]+)"', page.text)
    assert match is not None
    return match.group(1)


# ---------------------------------------------------------------------------
# Workbook builder — minimal Excel-import workbook, two investments
# ---------------------------------------------------------------------------


def _build_minimal_v2_workbook() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_names = [
        "Attributes",
        "Cash Flow In actual",
        "Cash Flow In plan",
        "Cash Flow Out actual",
        "Cash Flow Out plan",
        "NAVs actual",
        "NAVs plan",
        "total return actual",
        "total return plan",
        "interest rates",
    ]
    for n in sheet_names:
        wb.create_sheet(n)

    ws = wb["Attributes"]
    ws.append([None, "Investition A", "Investition B"])
    ws.append([None, "Aktien", "Private Equity"])
    ws.append([None, "Large Cap", "Buyout"])
    ws.append(["Region", "Europa", "USA"])
    ws.append(["Asset Class", "listed_equity", "private_equity"])
    ws.append(["Manager / Fondsname", "GP A", "GP B"])
    ws.append(["Vintage Year", 2020, 2021])
    ws.append(["Währung", "EUR", "EUR"])

    d = datetime.datetime(2024, 1, 1)
    ws = wb["NAVs actual"]
    ws.append([None, "Investition A", "Investition B"])
    ws.append([None, "Aktien", "Private Equity"])
    ws.append([None, "Large Cap", "Buyout"])
    ws.append([d, 100, 200])

    ws = wb["Cash Flow Out actual"]
    ws.append([None, "Investition A", "Investition B"])
    ws.append([None, "Aktien", "Private Equity"])
    ws.append([None, "Large Cap", "Buyout"])
    ws.append([d, -50, -75])

    for n in (
        "Cash Flow In actual",
        "Cash Flow In plan",
        "Cash Flow Out plan",
        "NAVs plan",
        "total return actual",
        "total return plan",
    ):
        ws = wb[n]
        ws.append([None, "Investition A", "Investition B"])
        ws.append([None, "Aktien", "Private Equity"])
        ws.append([None, "Large Cap", "Buyout"])
        ws.append([d, None, None])

    ws = wb["interest rates"]
    ws.append([None, "risk free rate"])
    ws.append([None, None])
    ws.append([None, None])
    ws.append([d, 0.04])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _upload_workbook(client: AsyncClient, csrf: str, filename: str = "x.xlsx") -> str:
    # 6F-5: the standalone /data-import POST route was sunset in favour
    # of the HTMX section upload endpoint. Successful uploads return
    # 200 with the Stage 2 fragment as body (not a 303 redirect).
    response = await client.post(
        "/api/data-import/section/upload",
        data={"csrf_token": csrf},
        files={
            "file": (
                filename,
                _build_minimal_v2_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        headers={"HX-Request": "true"},
        follow_redirects=False,
    )
    assert response.status_code == 200, response.text
    return filename


async def _resolve_upload_id(fresh_superuser_engine: AsyncEngine, filename: str) -> UUID:
    async with fresh_superuser_engine.connect() as conn:
        result = await conn.execute(
            text(
                "SELECT id FROM data_uploads WHERE filename = :fn ORDER BY created_at DESC LIMIT 1"
            ),
            {"fn": filename},
        )
        row = result.first()
    assert row is not None, f"No upload row found for {filename!r}"
    return UUID(str(row[0]))


# ---------------------------------------------------------------------------
# IRT-01: CSRF protection
# ---------------------------------------------------------------------------


async def test_irt01_post_without_csrf_returns_403(
    web_client: AsyncClient,
    seeded_user_and_classes: tuple[UUID, str, str],
    fresh_superuser_engine: AsyncEngine,
) -> None:
    _id, email, password = seeded_user_and_classes
    csrf = await _login_and_get_csrf(web_client, email, password)
    fname = await _upload_workbook(web_client, csrf, filename="csrf.xlsx")
    upload_id = await _resolve_upload_id(fresh_superuser_engine, fname)

    response = await web_client.post(
        f"/api/data-uploads/{upload_id}/import-as-investments",
        # No CSRF header.
        follow_redirects=False,
    )
    assert response.status_code == 403


# ---------------------------------------------------------------------------
# IRT-02: dry-run returns counts but writes nothing
# ---------------------------------------------------------------------------


async def test_irt02_dry_run_returns_counts_no_writes(
    web_client: AsyncClient,
    seeded_user_and_classes: tuple[UUID, str, str],
    fresh_superuser_engine: AsyncEngine,
) -> None:
    _id, email, password = seeded_user_and_classes
    csrf = await _login_and_get_csrf(web_client, email, password)
    fname = await _upload_workbook(web_client, csrf, filename="dry.xlsx")
    upload_id = await _resolve_upload_id(fresh_superuser_engine, fname)

    response = await web_client.post(
        f"/api/data-uploads/{upload_id}/import-as-investments?dry_run=true",
        headers={"X-CSRF-Token": csrf},
        follow_redirects=False,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["dry_run"] is True
    assert body["investments_created"] == 2
    assert body["navs_replaced"] == 2

    # No rows landed.
    async with fresh_superuser_engine.connect() as conn:
        result = await conn.execute(
            text("SELECT COUNT(*) FROM investments WHERE tenant_id = :tid"),
            {"tid": str(SENTINEL_TENANT_ID)},
        )
        count = result.scalar_one()
    assert count == 0


# ---------------------------------------------------------------------------
# IRT-03: real run writes the normalised rows
# ---------------------------------------------------------------------------


async def test_irt03_real_run_writes_normalised_rows(
    web_client: AsyncClient,
    seeded_user_and_classes: tuple[UUID, str, str],
    fresh_superuser_engine: AsyncEngine,
) -> None:
    _id, email, password = seeded_user_and_classes
    csrf = await _login_and_get_csrf(web_client, email, password)
    fname = await _upload_workbook(web_client, csrf, filename="real.xlsx")
    upload_id = await _resolve_upload_id(fresh_superuser_engine, fname)

    response = await web_client.post(
        f"/api/data-uploads/{upload_id}/import-as-investments",
        headers={"X-CSRF-Token": csrf},
        follow_redirects=False,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["dry_run"] is False
    assert body["investments_created"] == 2

    async with fresh_superuser_engine.connect() as conn:
        invs = await conn.execute(
            text(
                "SELECT name, investment_type FROM investments WHERE tenant_id = :tid ORDER BY name"
            ),
            {"tid": str(SENTINEL_TENANT_ID)},
        )
        names_and_types = [(r[0], r[1]) for r in invs.all()]
    assert names_and_types == [
        ("Investition A", "listed_equity"),
        ("Investition B", "private_equity"),
    ]


# ---------------------------------------------------------------------------
# IRT-04: 404 on unknown / foreign-tenant upload id
# ---------------------------------------------------------------------------


async def test_irt04_unknown_upload_id_returns_404(
    web_client: AsyncClient,
    seeded_user_and_classes: tuple[UUID, str, str],
) -> None:
    _id, email, password = seeded_user_and_classes
    csrf = await _login_and_get_csrf(web_client, email, password)

    response = await web_client.post(
        f"/api/data-uploads/{uuid4()}/import-as-investments",
        headers={"X-CSRF-Token": csrf},
        follow_redirects=False,
    )
    assert response.status_code == 404


# ---------------------------------------------------------------------------
# IRT-05: malformed JSONB snapshot → 400 (ImportFormatError translation)
# ---------------------------------------------------------------------------


async def test_irt05_missing_attributes_sheet_returns_400(
    web_client: AsyncClient,
    seeded_user_and_classes: tuple[UUID, str, str],
    fresh_superuser_engine: AsyncEngine,
) -> None:
    """Route translates ``ImportFormatError`` to a 400 with a message body.

    Phase-2 upload validation accepts the workbook (it has the
    ``Attributes`` sheet) but a downstream operation could plausibly
    leave the JSONB snapshot inconsistent — e.g. a future maintenance
    task that drops a sheet or a DB-level corruption. The contract
    here is that the route surfaces the structural failure as 400 (a
    client-correctable condition) rather than 500.

    We provoke the structural failure deterministically by deleting
    the ``attributes`` row from ``data_upload_sheets`` after upload,
    then drive the import endpoint and assert the 400 + message body.
    """
    _id, email, password = seeded_user_and_classes
    csrf = await _login_and_get_csrf(web_client, email, password)
    fname = await _upload_workbook(web_client, csrf, filename="malformed.xlsx")
    upload_id = await _resolve_upload_id(fresh_superuser_engine, fname)

    # Drop the Attributes sheet from the JSONB snapshot.
    async with fresh_superuser_engine.begin() as conn:
        await conn.execute(
            text(
                "DELETE FROM data_upload_sheets "
                "WHERE upload_id = :uid AND sheet_name = 'attributes'"
            ),
            {"uid": str(upload_id)},
        )

    response = await web_client.post(
        f"/api/data-uploads/{upload_id}/import-as-investments",
        headers={"X-CSRF-Token": csrf},
        follow_redirects=False,
    )
    assert response.status_code == 400, response.text
    body = response.json()
    # FastAPI wraps HTTPException details under "detail".
    assert "Attributes" in body["detail"] or "attributes" in body["detail"]
